import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl import load_workbook
from assets.func.uteis.popUp import popUp

arquivo_selecionado = None

def selecionar_arquivo():
    """Abre um diálogo para seleção de arquivo e armazena o caminho do arquivo selecionado."""
    global arquivo_selecionado  
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal
    arquivo_selecionado = filedialog.askopenfilename(title="Escolha o arquivo da planilha", filetypes=[("Arquivos Excel", "*.xlsx")])

    if not arquivo_selecionado:
        print("Nenhum arquivo selecionado")
        popUp("Nenhuma planilha foi selecionada.")  # Mostra uma mensagem na interface
        return None  # Retorna None para indicar erro

    return arquivo_selecionado

def info_planilha(index_pagina):
    """Carrega a página da planilha com base no índice informado."""
    global arquivo_selecionado
    
    if arquivo_selecionado is None:
        if selecionar_arquivo() is None:  # Tenta selecionar e verifica se deu certo
            return None

    if index_pagina == -1:
        print("Nenhuma página selecionada")
        popUp("Nenhuma página foi selecionada.")
        return None
    
    try:
        planilha = load_workbook(arquivo_selecionado)
        nome_pagina = planilha.sheetnames[index_pagina]
        pagina = planilha[nome_pagina]
        return pagina
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        popUp("Erro ao carregar a planilha. Verifique o arquivo.")
        return None

def listar_paginas():
    """Retorna uma lista com os nomes das páginas da planilha."""
    global arquivo_selecionado

    if arquivo_selecionado is None:
        selecionar_arquivo()
            

    try:
        planilha = openpyxl.load_workbook(arquivo_selecionado, read_only=True)
        nomes_paginas = planilha.sheetnames
        planilha.close()
        return nomes_paginas
    except FileNotFoundError:
        print("Erro: Arquivo não encontrado.")
        popUp("Arquivo não encontrado.")
        return []
    except Exception as e:
        print(f"Erro inesperado: {e}")        
        return []
