import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl import load_workbook
from assets.func.uteis.popUp import popUp


arquivo_selecionado = None

def selecionar_arquivo():
    global arquivo_selecionado  # Torna a variável global
    root = tk.Tk()
    root.withdraw()  # Oculta a janela principal
    arquivo_selecionado = filedialog.askopenfilename(title="Escolha o arquivo da planilha", filetypes=[("Arquivos Excel", "*.xlsx")])
    #print(arquivo_selecionado)

    if not arquivo_selecionado:
        print("Nenhum arquivo selecionado")
        #popUp("Nenhum arquivo selecionado")
        return False
    

def info_planilha(index_pagina):
    global arquivo_selecionado
    
    if arquivo_selecionado is None:
        selecionar_arquivo()

    if index_pagina == -1:
        print("Nenhuma página selecionada")
        #popUp("Nenhuma página selecionada")
        return False
    
    planilha = load_workbook(arquivo_selecionado)
    nome_pagina = planilha.sheetnames[index_pagina]
    pagina = planilha[nome_pagina]

    return pagina

def listar_paginas():
    global arquivo_selecionado

    if arquivo_selecionado is None:
        selecionar_arquivo()

    try:
        # Abre a planilha em modo somente leitura        
        planilha = openpyxl.load_workbook(arquivo_selecionado, read_only=True)
        
        # Pega os nomes das sheets
        nomes_paginas = planilha.sheetnames      
        planilha.close()
        
        #print(f'nomes das paginas: {nomes_paginas}')
        return nomes_paginas
    
    except FileNotFoundError:
        print("Erro: Arquivo não encontrado. Verifique o caminho do arquivo.")
        return []
    except Exception as e:
        print(f"Erro inesperado ao listar: {e}")
        return []
